#!/usr/bin/env python3
"""Export an explicit allowlist of observation sheets from the IDR workbooks."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROBLEM = ROOT / "problems_complete" / "s41589-026-02251-9_idr-condensate-serine-charge"
SOURCE = PROBLEM / "data" / "nature-supplementary"
DESTINATION = PROBLEM / "curated" / "observations"
HOLDOUT_SOURCE = PROBLEM / "evaluator" / "holdout_source"
HOLDOUT_FILES = {
    "aromatic_serine_mutant_scores.csv",
    "charged_serine_mutant_scores.csv",
    "low_charge_mutant_scores.csv",
}

SHEETS = {
    "41589_2026_2251_MOESM10_ESM.xlsx": {
        "IDR_Pair_Miscibility(Fig1b)": "idr_pair_miscibility.csv",
    },
    "41589_2026_2251_MOESM11_ESM.xlsx": {
        "S_PairScores(Fig2c)": "serine_pair_scores.csv",
        "FYW_PairScores(Fig2f)": "aromatic_pair_scores.csv",
        "S_FYW_4grp_PairScores(Fig2h)": "serine_aromatic_group_pair_scores.csv",
    },
    "41589_2026_2251_MOESM12_ESM.xlsx": {
        "EDKRH_PairScores(Fig3c)": "charged_pair_scores.csv",
        "EDKRH_S_4grp_PairScores(Fig3h)": "charged_serine_group_pair_scores.csv",
    },
    "41589_2026_2251_MOESM14_ESM.xlsx": {
        "SaPS_WT_vs_Phos_Data(Fig5b)": "saps_phosphorylation_pairs.csv",
        "SY_Frac_vs_PhosLevelData(Fig5d)": "serine_tyrosine_phosphorylation.csv",
        "SRSF1_PhosphoMimic_vitro(Fig5g)": "srsf1_phosphomimic_in_vitro.csv",
        "SRSF1_CLK1_vitro(Fig5i)": "srsf1_clk1_in_vitro.csv",
        "SRSF1_PhosphoMimic_vivo(Fig5k)": "srsf1_phosphomimic_in_vivo.csv",
    },
    "41589_2026_2251_MOESM15_ESM.xlsx": {
        "Dual-Luc-Assay(Fig6a)": "dual_luciferase_assay.csv",
        "Input_21IDRs(Fig6b,c,d)": "idr21_features.csv",
        "Input_SOX2_r_wt-mut(Fig6k)": "sox2_mutant_inputs.csv",
        "qRT-PCR_Results(Fig6l)": "qrt_pcr_measurements.csv",
    },
    "41589_2026_2251_MOESM17_ESM.xlsx": {
        "Miscibility_inCellinVitro(EDF2)": "miscibility_cell_vitro.csv",
    },
    "41589_2026_2251_MOESM21_ESM.xlsx": {
        "NCPR_PairScores(EDF6f)": "ncpr_pair_scores.csv",
        "ChargePat_Pairs(EDF6i)": "charge_pattern_pairs.csv",
    },
    "41589_2026_2251_MOESM26_ESM.xlsx": {
        "Charge_vs_WT_Luc(EDF10i)": "charge_vs_wt_luciferase.csv",
    },
    "41589_2026_2251_MOESM9_ESM.xlsx": {
        "BbSwap_PairScores(SIFig5b)": "backbone_swap_pair_scores.csv",
        "BackboneSwap_45pairs(SIFig5e)": "backbone_swap_inputs.csv",
        "FYWrich_HighSFus_Sc(SIFig7a)": "aromatic_serine_fusion_scores.csv",
        "FYWrich_SplusMut_Sc(SIFig7b)": "aromatic_serine_mutant_scores.csv",
        "SRich_Sblocky_Scores(SIFig8de)": "serine_blocky_high_scores.csv",
        "SPoor_Sblocky_Scores(SIFig8ij)": "serine_blocky_low_scores.csv",
        "ChargHigh_HighSFus_Sc(SIFig9c)": "charged_serine_fusion_scores.csv",
        "ChargHigh_SplusMut_Sc(SIFig9g)": "charged_serine_mutant_scores.csv",
        "ChargLow_CRichFus_Sc(SIFig10a)": "low_charge_fusion_scores.csv",
        "ChargLow_ChdMut_Sc(SIFig10b)": "low_charge_mutant_scores.csv",
        "PolII_MiscibVsTA_Data(SIFig14b)": "polii_miscibility_transactivation.csv",
        "OCT4Nmyc_Miscib(SIFig15ek)": "oct4_nmyc_miscibility.csv",
        "OCT4Nmyc_qPCR_Data(SIFig15fl)": "oct4_nmyc_qpcr.csv",
    },
}


def main() -> int:
    DESTINATION.mkdir(parents=True, exist_ok=True)
    HOLDOUT_SOURCE.mkdir(parents=True, exist_ok=True)
    expected = {name for sheets in SHEETS.values() for name in sheets.values()}
    for stale in DESTINATION.glob("*.csv"):
        if stale.name not in expected or stale.name in HOLDOUT_FILES:
            stale.unlink()
    for workbook_name, sheets in SHEETS.items():
        workbook = load_workbook(
            SOURCE / workbook_name,
            read_only=True,
            data_only=True,
        )
        for sheet_name, output_name in sheets.items():
            worksheet = workbook[sheet_name]
            output_root = HOLDOUT_SOURCE if output_name in HOLDOUT_FILES else DESTINATION
            with (output_root / output_name).open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                for row in worksheet.iter_rows(values_only=True):
                    values = list(row)
                    while values and values[-1] is None:
                        values.pop()
                    writer.writerow("" if value is None else value for value in values)
    print(f"Wrote {len(expected)} neutral observation tables to {DESTINATION}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
